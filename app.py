import datetime
import tkinter as tk
from tkinter import ttk
import openpyxl
from tkcalendar import Calendar
from tkinter import filedialog, messagebox
from search import *
import threading

root = tk.Tk()
root.title("高校理论文章检索器")
root.configure(bg="white")
root.geometry("550x300")
root.resizable(False, False)

isRunning = False

def get_selected_date():
    if isRunning:
        messagebox.showinfo("提示", "正在检索，请稍等！")
        return
    if not cal.get_date() or not cal2.get_date():
        messagebox.showinfo("提示", "请选择日期！")
        return
    if not path_label.cget("text"):
        messagebox.showinfo("提示", "请选择路径！")
        return
    start_date:str = cal.get_date()
    end_data:str= cal2.get_date()
    if start_date > end_data:
        messagebox.showinfo("提示", "起始日期不能大于结束日期！")
        return
    
    dayNum = (datetime.datetime.strptime(end_data, "%Y-%m-%d") - datetime.datetime.strptime(start_date, "%Y-%m-%d")).days+1
    if dayNum > 30:
        messagebox.showinfo("提示", "日期范围不能超过30天！")
        return
    
    # 进度条窗口
    progress_dialog = tk.Toplevel(root)
    progress_dialog.geometry("300x100")
    # progress_dialog.wait_window()
    progress_dialog.title("检索进度")
    progress_label = tk.Label(progress_dialog, text="正在检索...")
    progress_label.pack(padx=20, pady=10)
    progress_bar = ttk.Progressbar(progress_dialog, orient="horizontal", length=200)
    progress_bar.pack(padx=20, pady=10)

    detail={
        "totaldays":0,
        "currentday":0,
        "currentprogress":0
    }
    def stepProgress():
        stepRatio=(detail["currentprogress"]+4*detail["currentday"]) / (4*detail["totaldays"])*100
        progress_bar["value"] = int(stepRatio)
        
        if detail["currentprogress"]==1:
            progress_label.config(text="正在检索第{}天光明日报...".format(detail["currentday"]+1))
        elif detail["currentprogress"]==2:
            progress_label.config(text="正在检索第{}天解放日报...".format(detail["currentday"]+1))
        elif detail["currentprogress"]==3:
            progress_label.config(text="正在检索第{}天文汇报...".format(detail["currentday"]+1))
        elif detail["currentprogress"]==4:
            progress_label.config(text="正在检索第{}天人民日报...".format(detail["currentday"]+2))
        
        # progress_bar["value"] += 25
        # if progress_bar["value"] == 25:
        #     progress_label.config(text="正在检索光明日报...")
        # elif progress_bar["value"] == 50:
        #     progress_label.config(text="正在检索解放日报...")
        # elif progress_bar["value"] == 75:
        #     progress_label.config(text="正在检索文汇报...")
    def run():
        global isRunning
        isRunning = True
        wb=openpyxl.Workbook()
        wb.remove(wb.active)
        wsRMRB=wb.create_sheet("人民日报")
        wsRMRB.append(["报刊名","日期","具体版面","文章名称","作者","具体所属单位","字数","链接"])
        wsGMRB=wb.create_sheet("光明日报")
        wsGMRB.append(["报刊名","日期","具体版面","文章名称","作者","具体所属单位","字数","链接"])
        wsJFRB=wb.create_sheet("解放日报")
        wsJFRB.append(["报刊名","日期","具体版面","文章名称","作者","具体所属单位","字数","链接"])
        wsWHB=wb.create_sheet("文汇报")
        wsWHB.append(["报刊名","日期","具体版面","文章名称","作者","具体所属单位","字数","链接"])
        ps=wb.create_sheet("潜在文章")
        ps.append(["报刊名","日期","具体版面","文章名称""链接"])
        
        for day in range(dayNum):
            date=datetime.datetime.strptime(start_date,"%Y-%m-%d")+datetime.timedelta(days=day)
            split_date = str(date).split(" ")[0].split("-")
            year = split_date[0]
            month = split_date[1]
            date = split_date[2]
            
            detail["totaldays"]=dayNum
            detail["currentday"]=day
            detail["currentprogress"]=1
            search_article_RMRB(year, month, date,wsRMRB,ps, stepProgress)
            detail["currentprogress"]=2
            search_article_GMRB(year, month, date,wsGMRB,ps, stepProgress)
            detail["currentprogress"]=3
            search_article_JFRB(year, month, date,wsJFRB,ps, stepProgress)
            detail["currentprogress"]=4
            search_article_WHB(year, month, date,wsWHB,ps, stepProgress)
        
        timestamp=datetime.datetime.now().timestamp()
        start=start_date.split("-")
        end=end_data.split("-")
        wb.save(path_label.cget("text")+"/高校理论文章检索结果-{}-{}-{}-to-{}-{}-{}-timestamp-{}.xlsx".format(start[0],start[1],start[2],end[0],end[1],end[2],timestamp))
        isRunning = False
        progress_bar.stop()
        progress_dialog.destroy()
        messagebox.showinfo("提示", "检索完成！")
            
    calculation_thread = threading.Thread(target=run)
    calculation_thread.start()
    progress_dialog.wait_window()
    

def choose_directory():
    directory_path = filedialog.askdirectory()
    if directory_path:
        path_label.config(text=directory_path)

callabel = tk.Label(root, text="起始日期")
callabel.grid(row=0, column=0)

callabel2 = tk.Label(root, text="结束日期")
callabel2.grid(row=0, column=1)

cal = Calendar(root, selectmode="day", date_pattern="y-mm-dd")
cal.grid(row=1, column=0,padx=10, pady=10)

cal2 = Calendar(root, selectmode="day", date_pattern="y-mm-dd")
cal2.grid(row=1, column=1,padx=10, pady=10)

btn = tk.Button(root, text="选择路径", command=choose_directory)
btn.grid(row=2, column=0)

# 创建标签，用于显示选择的路径
path_label = tk.Label(root, text="")
path_label.grid(row=2, column=1)

btn = ttk.Button(root, text="检索指定日期的理论文章", command=get_selected_date)
btn.grid(row=3, column=0, columnspan=2, pady=10)

def main():
    root.mainloop()

if __name__ == "__main__":
    main()